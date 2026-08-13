"""F7/F8:純記憶體檔案預覽、批次脫敏與稽核報告。"""

import base64
import binascii
import re
from collections import defaultdict
import csv
from datetime import datetime, timezone
from io import BytesIO, StringIO
import json
from pathlib import Path
from typing import Any
from zipfile import ZIP_DEFLATED, BadZipFile, ZipFile

import anyio
from fastapi import HTTPException, UploadFile
from openpyxl import Workbook, load_workbook
from starlette.formparsers import MultiPartParser

from app import config
from app.engine import (
    BatchTimeoutError,
    EntitySelectionError,
    ScrubEngine,
    TextTooLongError,
)
from app.recognizers.ckip_ner import CkipTimeoutError


MAX_FILE_SIZE = 10 * 1024 * 1024
SUPPORTED_FILETYPES = {"txt", "csv", "xlsx"}
VALID_COLUMN_STRATEGIES = {"mask_all", "ner", "skip"}

# Starlette 預設超過 1 MiB 會把 multipart file rollover 到磁碟。
# max_size=0 的 SpooledTemporaryFile 永不自動 rollover，符合全程記憶體鐵律。
MultiPartParser.spool_max_size = 0


def _unique_zip_name(raw_name: str, occupied: set[str]) -> str:
    # 消毒須同時涵蓋 POSIX 與 Windows 規則:在 macOS/Linux 上
    # Path().name 不會把反斜線視為分隔符,「..\\evil.txt」會原樣進 ZIP
    basename = raw_name.replace("\\", "/").split("/")[-1]
    basename = re.sub(r"^[A-Za-z]:", "", basename)
    basename = "".join(
        c for c in basename if c.isprintable() and c not in '<>:"|?*'
    ).strip(" .")
    if basename in {"", ".", ".."}:
        basename = "file"
    candidate = basename
    suffix = Path(basename).suffix
    stem = basename[: -len(suffix)] if suffix else basename
    duplicate_index = 2
    while candidate in occupied:
        candidate = f"{stem} ({duplicate_index}){suffix}"
        duplicate_index += 1
    occupied.add(candidate)
    return candidate


def bundle_zip(entries: Any) -> dict:
    if not isinstance(entries, list):
        raise HTTPException(status_code=400, detail="entries 必須是陣列")
    if len(entries) > config.MAX_ZIP_ENTRIES:
        raise HTTPException(
            status_code=400,
            detail=f"ZIP 項目不可超過 {config.MAX_ZIP_ENTRIES} 筆",
        )

    # 先以編碼長度估算總量(b64 → bytes 約 3/4),超限即拒絕,
    # 避免先做大量 decode 配置記憶體才發現超標
    estimated_total = 0
    for entry in entries:
        if not isinstance(entry, dict):
            raise HTTPException(status_code=400, detail="ZIP 項目格式錯誤")
        if not isinstance(entry.get("filename"), str) or not isinstance(
            entry.get("content_b64"), str
        ):
            raise HTTPException(status_code=400, detail="ZIP 項目格式錯誤")
        estimated_total += (len(entry["content_b64"]) * 3) // 4
        if estimated_total > config.MAX_ZIP_TOTAL_BYTES:
            raise HTTPException(status_code=413, detail="ZIP 解碼後總量不可超過 60MB")

    occupied: set[str] = set()
    decoded_entries: list[tuple[str, bytes]] = []
    total_bytes = 0
    for entry in entries:
        raw_name = entry["filename"]
        content_b64 = entry["content_b64"]
        try:
            content = base64.b64decode(content_b64, validate=True)
        except (binascii.Error, ValueError) as exc:
            raise HTTPException(status_code=400, detail="content_b64 不是有效 Base64") from exc
        total_bytes += len(content)
        if total_bytes > config.MAX_ZIP_TOTAL_BYTES:
            raise HTTPException(status_code=413, detail="ZIP 解碼後總量不可超過 60MB")
        decoded_entries.append((_unique_zip_name(raw_name, occupied), content))

    output = BytesIO()
    with ZipFile(output, mode="w", compression=ZIP_DEFLATED) as archive:
        for filename, content in decoded_entries:
            archive.writestr(filename, content)

    timestamp = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    return {
        "zip_b64": base64.b64encode(output.getvalue()).decode("ascii"),
        "filename": f"tw-pii-scrubbed-{timestamp}.zip",
    }


def _filetype(filename: str | None) -> str:
    suffix = Path(filename or "").suffix.lower().lstrip(".")
    if suffix not in SUPPORTED_FILETYPES:
        raise HTTPException(status_code=415, detail="僅支援 .txt、.csv、.xlsx 檔案")
    return suffix


async def _read_upload(file: UploadFile) -> tuple[str, str, BytesIO]:
    filetype = _filetype(file.filename)
    content = await file.read(MAX_FILE_SIZE + 1)
    if len(content) > MAX_FILE_SIZE:
        raise HTTPException(status_code=413, detail="檔案大小不可超過 10MB")
    return file.filename or "", filetype, BytesIO(content)


def _decode_text(content: bytes) -> tuple[str, str]:
    for codec, reported_encoding in (
        ("utf-8-sig", "utf-8"),
        ("utf-8", "utf-8"),
        ("big5", "big5"),
        ("cp950", "big5"),
    ):
        try:
            return content.decode(codec), reported_encoding
        except UnicodeDecodeError:
            continue
    raise HTTPException(status_code=400, detail="檔案無法解碼，請轉存為 UTF-8 後重試")


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def _parse_csv(buffer: BytesIO) -> tuple[list[str], list[list[str]], str]:
    text, encoding = _decode_text(buffer.getvalue())
    try:
        rows = list(csv.reader(StringIO(text, newline="")))
    except csv.Error as exc:
        raise HTTPException(status_code=400, detail=f"CSV 格式無法解析:{exc}") from exc
    if not rows:
        return [], [], encoding
    return rows[0], rows[1:], encoding


def _parse_xlsx(buffer: BytesIO) -> tuple[list[str], list[list[str]], str]:
    # 解析前先驗 OOXML ZIP 結構:10MB 的壓縮檔可能解壓成數 GB
    # (zip bomb),必須在 openpyxl materialize 之前擋下
    buffer.seek(0)
    try:
        with ZipFile(buffer) as zf:
            infos = zf.infolist()
    except BadZipFile as exc:
        raise HTTPException(status_code=400, detail="XLSX 格式無法解析") from exc
    if len(infos) > config.MAX_XLSX_ZIP_ENTRIES:
        raise HTTPException(status_code=400, detail="XLSX 內部項目數異常")
    if sum(info.file_size for info in infos) > config.MAX_XLSX_UNCOMPRESSED_BYTES:
        raise HTTPException(status_code=400, detail="XLSX 解壓後大小超過限制")

    buffer.seek(0)
    try:
        workbook = load_workbook(buffer, read_only=True, data_only=False)
        try:
            worksheet = workbook.worksheets[0]
            rows = []
            for row_index, row in enumerate(worksheet.iter_rows(values_only=True)):
                if row_index >= config.MAX_XLSX_ROWS:
                    raise HTTPException(
                        status_code=400,
                        detail=f"列數超過上限({config.MAX_XLSX_ROWS:,} 列)",
                    )
                if row and len(row) > config.MAX_XLSX_COLS:
                    raise HTTPException(
                        status_code=400,
                        detail=f"欄數超過上限({config.MAX_XLSX_COLS} 欄)",
                    )
                rows.append([_cell_text(value) for value in row])
            title = worksheet.title
        finally:
            workbook.close()
    except HTTPException:
        raise
    except Exception as exc:
        # openpyxl 對不同壞檔可能拋 BadZipFile、XML ParseError 或 ValueError。
        raise HTTPException(status_code=400, detail="XLSX 格式無法解析") from exc
    if not rows:
        return [], [], title
    return rows[0], rows[1:], title


def _column_preview(headers: list[str], rows: list[list[str]]) -> list[dict]:
    columns = []
    for column_index, header in enumerate(headers):
        samples = []
        for row in rows:
            value = row[column_index] if column_index < len(row) else ""
            if value != "":
                samples.append(value)
            if len(samples) == 3:
                break
        columns.append({"name": header, "samples": samples})
    return columns


async def preview_file(file: UploadFile) -> dict:
    filename, filetype, buffer = await _read_upload(file)
    # 解析(xlsx 可能較慢)卸載到 worker thread,避免阻塞 event loop
    return await anyio.to_thread.run_sync(_preview_sync, filename, filetype, buffer)


def _preview_sync(filename: str, filetype: str, buffer: BytesIO) -> dict:
    if filetype == "txt":
        text, encoding = _decode_text(buffer.getvalue())
        return {
            "filename": filename,
            "filetype": filetype,
            "encoding": encoding,
            "columns": None,
            "row_count": len(text.splitlines()),
        }
    if filetype == "csv":
        headers, rows, encoding = _parse_csv(buffer)
    else:
        headers, rows, _sheet_title = _parse_xlsx(buffer)
        encoding = None
    return {
        "filename": filename,
        "filetype": filetype,
        "encoding": encoding,
        "columns": _column_preview(headers, rows),
        "row_count": len(rows),
    }


def _parse_entities(raw: str) -> list[str] | None:
    try:
        entities = json.loads(raw)
    except (json.JSONDecodeError, TypeError) as exc:
        raise HTTPException(status_code=400, detail="entities 必須是 JSON 陣列") from exc
    if entities is None:
        return None
    if not isinstance(entities, list) or not all(
        isinstance(entity, str) for entity in entities
    ):
        raise HTTPException(status_code=400, detail="entities 必須是 JSON 字串陣列")
    return entities


def _parse_column_strategies(raw: str, headers: list[str]) -> dict[str, str]:
    try:
        strategies = json.loads(raw)
    except (json.JSONDecodeError, TypeError) as exc:
        raise HTTPException(
            status_code=400, detail="column_strategies 必須是 JSON 物件"
        ) from exc
    if not isinstance(strategies, dict) or not all(
        isinstance(name, str) and isinstance(strategy, str)
        for name, strategy in strategies.items()
    ):
        raise HTTPException(
            status_code=400, detail="column_strategies 必須是 JSON 字串對應物件"
        )
    invalid = [
        name for name, strategy in strategies.items() if strategy not in VALID_COLUMN_STRATEGIES
    ]
    if invalid:
        raise HTTPException(
            status_code=400,
            detail=f"欄位策略無效:{', '.join(invalid)}",
        )
    missing = list(dict.fromkeys(name for name in headers if name not in strategies))
    if missing:
        raise HTTPException(
            status_code=400,
            detail=f"缺少欄位策略:{', '.join(missing)}",
        )
    return strategies


def _mask_column(value: str) -> str:
    return value[0] + "○" * (len(value) - 1) if value else value


class _ColumnPlaceholderAllocator:
    """mask_all 專用:每欄同值共用編號，且不覆蓋 NER mapping。"""

    def __init__(self, occupied: set[str]) -> None:
        self._occupied = occupied
        self._counters: dict[str, int] = defaultdict(int)
        self._by_value: dict[str, dict[str, str]] = defaultdict(dict)
        self.mapping: dict[str, str] = {}

    def get(self, column: str, value: str) -> str:
        if value in self._by_value[column]:
            return self._by_value[column][value]
        while True:
            self._counters[column] += 1
            placeholder = f"<{column}_{self._counters[column]}>"
            if placeholder not in self._occupied:
                break
        self._occupied.add(placeholder)
        self._by_value[column][value] = placeholder
        self.mapping[placeholder] = value
        return placeholder


def _report_finding(finding: dict, location: dict) -> dict:
    return {
        "entity_type": finding["entity_type"],
        "original_masked": finding["original_masked"],
        "score": finding["score"],
        "recognizer": finding["recognizer"],
        "replacement": finding["replacement"],
        "location": location,
    }


def _make_report(
    *,
    filename: str,
    filetype: str,
    mode: str,
    column_strategies: dict[str, str] | None,
    stats: dict[str, int],
    findings: list[dict],
) -> dict:
    return {
        "timestamp": datetime.now(timezone.utc).isoformat(),
        "tool": "TW-PII-Scrubber",
        "source_filename": filename,
        "filetype": filetype,
        "mode": mode,
        "column_strategies": column_strategies,
        "stats": stats,
        "findings": findings,
    }


def _process_table(
    headers: list[str],
    rows: list[list[str]],
    strategies: dict[str, str],
    mode: str,
    entities: list[str] | None,
    engine: ScrubEngine,
) -> tuple[list[list[str]], dict[str, int], list[dict], dict[str, str]]:
    output_rows = [row.copy() for row in rows]
    ner_cells: list[tuple[int, int, str, str]] = []
    for row_index, row in enumerate(rows, start=2):
        if len(row) > len(headers):
            raise HTTPException(status_code=400, detail=f"第 {row_index} 列欄位數超過表頭")
        for column_index, column in enumerate(headers):
            if strategies[column] == "ner":
                value = row[column_index] if column_index < len(row) else ""
                ner_cells.append((row_index, column_index, column, value))

    stats: dict[str, int] = defaultdict(int)
    located_findings: list[tuple[int, int, int, dict]] = []
    mapping: dict[str, str] = {}
    if ner_cells:
        batch = engine.scrub_batch(
            [value for _row, _index, _column, value in ner_cells],
            mode,
            entities,
        )
        for entity_type, count in batch["stats"].items():
            stats[entity_type] += count
        if mode == "placeholder":
            mapping.update(batch.get("mapping", {}))
        for (row_number, column_index, column, _value), result in zip(
            ner_cells, batch["results"]
        ):
            if column_index < len(output_rows[row_number - 2]):
                output_rows[row_number - 2][column_index] = result["scrubbed_text"]
            for finding_index, finding in enumerate(result["findings"]):
                located_findings.append(
                    (
                        row_number,
                        column_index,
                        finding_index,
                        _report_finding(
                            finding, {"row": row_number, "column": column}
                        ),
                    )
                )

    allocator = _ColumnPlaceholderAllocator(set(mapping))
    for row_number, row in enumerate(rows, start=2):
        for column_index, column in enumerate(headers):
            if strategies[column] != "mask_all" or column_index >= len(row):
                continue
            value = row[column_index]
            if value == "":
                continue
            replacement = (
                _mask_column(value)
                if mode == "mask"
                else allocator.get(column, value)
            )
            output_rows[row_number - 2][column_index] = replacement
            stats["COLUMN_MASK"] += 1
            located_findings.append(
                (
                    row_number,
                    column_index,
                    0,
                    {
                        "entity_type": "COLUMN_MASK",
                        "original_masked": _mask_column(value),
                        "score": 1.0,
                        "recognizer": "ColumnStrategy",
                        "replacement": replacement,
                        "location": {"row": row_number, "column": column},
                    },
                )
            )
    mapping.update(allocator.mapping)
    located_findings.sort(key=lambda item: item[:3])
    return output_rows, dict(stats), [item[3] for item in located_findings], mapping


def _csv_output(headers: list[str], rows: list[list[str]]) -> bytes:
    stream = StringIO(newline="")
    writer = csv.writer(stream)
    if headers:
        writer.writerow(headers)
    writer.writerows(rows)
    return stream.getvalue().encode("utf-8-sig")


def _xlsx_output(headers: list[str], rows: list[list[str]], title: str) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = title
    if headers:
        worksheet.append(headers)
    for row in rows:
        worksheet.append(row)
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def _mask_filename(filename: str, engine: ScrubEngine) -> str:
    """檔名本身可能含個資(王小明_A123456789.txt),進報告/輸出檔名/
    ZIP 前先過完整脫敏管線;遮罩字元 * 在 Windows 檔名不合法,改為 x。
    注意:NER 模型對孤立詞辨識較弱,檔名遮罩屬盡力而為(README 已註記)。"""
    stem = Path(filename).stem
    suffix = Path(filename).suffix
    try:
        masked = engine.scrub(stem, "mask")["scrubbed_text"]
    except Exception:
        masked = "file"  # fail-safe:遮不了就不透出原名
    masked = masked.replace("*", "x")
    masked = "".join(c for c in masked if c.isprintable() and c not in '<>:"/\\|?*')
    return (masked.strip() or "file") + suffix


def _output_filename(masked_source: str, filetype: str) -> str:
    return f"{Path(masked_source).stem}.scrubbed.{filetype}"


async def process_file(
    file: UploadFile,
    mode: str,
    entities_json: str,
    column_strategies_json: str,
    engine: ScrubEngine,
) -> dict:
    filename, filetype, buffer = await _read_upload(file)
    entities = _parse_entities(entities_json)
    # CPU 密集(CKIP 推論)卸載到 worker thread:async 端點內同步跑
    # 推論會卡死 event loop,處理大檔期間整個 GUI/健康檢查都會凍結
    return await anyio.to_thread.run_sync(
        _process_sync, filename, filetype, buffer, mode, entities,
        column_strategies_json, engine,
    )


def _process_sync(
    filename: str,
    filetype: str,
    buffer: BytesIO,
    mode: str,
    entities: list[str] | None,
    column_strategies_json: str,
    engine: ScrubEngine,
) -> dict:
    try:
        if filetype == "txt":
            text, _encoding = _decode_text(buffer.getvalue())
            result = engine.scrub(text, mode, entities)
            report_findings = [
                _report_finding(
                    finding,
                    {"start": finding["start"], "end": finding["end"]},
                )
                for finding in result["findings"]
            ]
            output_bytes = result["scrubbed_text"].encode("utf-8")
            stats = result["stats"]
            mapping = result.get("mapping", {})
            strategies = None
        else:
            if filetype == "csv":
                headers, rows, _encoding = _parse_csv(buffer)
                sheet_title = "Sheet"
            else:
                headers, rows, sheet_title = _parse_xlsx(buffer)
            strategies = _parse_column_strategies(column_strategies_json, headers)
            output_rows, stats, report_findings, mapping = _process_table(
                headers, rows, strategies, mode, entities, engine
            )
            output_bytes = (
                _csv_output(headers, output_rows)
                if filetype == "csv"
                else _xlsx_output(headers, output_rows, sheet_title)
            )
    except EntitySelectionError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except TextTooLongError as exc:
        raise HTTPException(status_code=413, detail=str(exc)) from exc
    except (BatchTimeoutError, CkipTimeoutError) as exc:
        raise HTTPException(
            status_code=408, detail="檔案處理逾時,請縮小檔案"
        ) from exc

    # 檔名可能含個資:輸出檔名與報告一律使用脫敏後的檔名
    masked_source = _mask_filename(filename, engine)
    response = {
        "filename": _output_filename(masked_source, filetype),
        "file_b64": base64.b64encode(output_bytes).decode("ascii"),
        "report": _make_report(
            filename=masked_source,
            filetype=filetype,
            mode=mode,
            column_strategies=strategies,
            stats=stats,
            findings=report_findings,
        ),
    }
    if mode == "placeholder":
        response["mapping"] = mapping
    return response
