"""F7/F8:檔案批次脫敏與稽核報告 API 測試。"""

import base64
import csv
from datetime import datetime
from io import BytesIO, StringIO
import json
from zipfile import ZipFile

from openpyxl import Workbook, load_workbook
import pytest
from fastapi.testclient import TestClient

from app.main import app


@pytest.fixture(scope="module")
def client():
    with TestClient(app, base_url="http://127.0.0.1:7860") as test_client:
        yield test_client


def _csv_bytes(rows: list[list[object]]) -> bytes:
    stream = StringIO(newline="")
    csv.writer(stream).writerows(rows)
    return stream.getvalue().encode("utf-8")


def _xlsx_bytes(tmp_path, rows: list[list[object]]) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    for row in rows:
        worksheet.append(row)
    path = tmp_path / "upload.xlsx"
    workbook.save(path)
    workbook.close()
    return path.read_bytes()


def _post_scrub_file(
    client,
    filename: str,
    content: bytes,
    *,
    mode: str = "mask",
    entities: list[str] | None = None,
    column_strategies: dict[str, str] | None = None,
):
    return client.post(
        "/api/scrub-file",
        files={"file": (filename, content, "application/octet-stream")},
        data={
            "mode": mode,
            "entities": json.dumps(entities, ensure_ascii=False),
            "column_strategies": json.dumps(
                column_strategies or {}, ensure_ascii=False
            ),
        },
    )


class TestFilePreview:
    @pytest.mark.parametrize(
        ("encoding", "expected_encoding"),
        [("utf-8", "utf-8"), ("big5", "big5")],
    )
    def test_txt_preview_reports_encoding_and_line_count(
        self, client, encoding: str, expected_encoding: str
    ) -> None:
        content = "第一行\n王小明\n第三行".encode(encoding)

        response = client.post(
            "/api/scrub-file/preview",
            files={"file": ("名單.txt", content, "text/plain")},
        )

        assert response.status_code == 200
        assert response.json() == {
            "filename": "名單.txt",
            "filetype": "txt",
            "encoding": expected_encoding,
            "columns": None,
            "row_count": 3,
        }

    def test_csv_preview_lists_nonempty_samples(self, client) -> None:
        content = _csv_bytes(
            [
                ["姓名", "備註", "編號"],
                ["王小明", "", "001"],
                ["李小華", "聯絡王小明", "002"],
                ["陳大文", "第三筆", "003"],
                ["第四人", "不應進樣本", "004"],
            ]
        )

        response = client.post(
            "/api/scrub-file/preview",
            files={"file": ("名單.csv", content, "text/csv")},
        )

        assert response.status_code == 200
        body = response.json()
        assert body["encoding"] == "utf-8"
        assert body["row_count"] == 4
        assert body["columns"] == [
            {"name": "姓名", "samples": ["王小明", "李小華", "陳大文"]},
            {"name": "備註", "samples": ["聯絡王小明", "第三筆", "不應進樣本"]},
            {"name": "編號", "samples": ["001", "002", "003"]},
        ]

    def test_xlsx_preview_has_null_encoding(self, client, tmp_path) -> None:
        content = _xlsx_bytes(
            tmp_path,
            [["姓名", "編號"], ["王小明", 1], ["李小華", 2]],
        )

        response = client.post(
            "/api/scrub-file/preview",
            files={
                "file": (
                    "名單.xlsx",
                    content,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )

        assert response.status_code == 200
        assert response.json() == {
            "filename": "名單.xlsx",
            "filetype": "xlsx",
            "encoding": None,
            "columns": [
                {"name": "姓名", "samples": ["王小明", "李小華"]},
                {"name": "編號", "samples": ["1", "2"]},
            ],
            "row_count": 2,
        }


class TestFileScrub:
    @pytest.mark.parametrize("encoding", ["utf-8", "big5"])
    def test_txt_upload_scrubs_utf8_and_big5(self, client, encoding: str) -> None:
        original = "本人王小明，身分證A123456789。"

        response = _post_scrub_file(
            client,
            "個資.txt",
            original.encode(encoding),
            entities=["PERSON", "TW_NATIONAL_ID"],
        )

        assert response.status_code == 200
        body = response.json()
        scrubbed = base64.b64decode(body["file_b64"]).decode("utf-8")
        assert body["filename"] == "個資.scrubbed.txt"
        assert "王○○" in scrubbed
        assert "A1******89" in scrubbed
        assert "mapping" not in body

    def test_csv_applies_mask_all_ner_and_skip_strategies(self, client) -> None:
        content = _csv_bytes(
            [
                ["姓名", "備註", "編號"],
                ["陳大文", "聯絡王小明", "A123456789"],
                ["王小明", "由王小明承辦", "KEEP-002"],
            ]
        )

        response = _post_scrub_file(
            client,
            "名單.csv",
            content,
            entities=["PERSON"],
            column_strategies={"姓名": "mask_all", "備註": "ner", "編號": "skip"},
        )

        assert response.status_code == 200
        body = response.json()
        raw_output = base64.b64decode(body["file_b64"])
        assert raw_output.startswith(b"\xef\xbb\xbf")
        rows = list(csv.reader(StringIO(raw_output.decode("utf-8-sig"))))
        assert rows == [
            ["姓名", "備註", "編號"],
            ["陳○○", "聯絡王○○", "A123456789"],
            ["王○○", "由王○○承辦", "KEEP-002"],
        ]
        assert body["report"]["stats"] == {"COLUMN_MASK": 2, "PERSON": 2}

    def test_placeholder_mode_reuses_numbers_and_returns_complete_mapping(
        self, client
    ) -> None:
        content = _csv_bytes(
            [
                ["姓名", "備註"],
                ["王小明", "王小明"],
                ["王小明", "王小明"],
            ]
        )

        response = _post_scrub_file(
            client,
            "名單.csv",
            content,
            mode="placeholder",
            entities=["PERSON"],
            column_strategies={"姓名": "mask_all", "備註": "ner"},
        )

        assert response.status_code == 200
        body = response.json()
        rows = list(
            csv.reader(
                StringIO(base64.b64decode(body["file_b64"]).decode("utf-8-sig"))
            )
        )
        assert rows[1:] == [
            ["<姓名_1>", "<PERSON_1>"],
            ["<姓名_1>", "<PERSON_1>"],
        ]
        assert body["mapping"] == {
            "<PERSON_1>": "王小明",
            "<姓名_1>": "王小明",
        }

    def test_xlsx_remains_valid_and_applies_all_strategies(
        self, client, tmp_path
    ) -> None:
        content = _xlsx_bytes(
            tmp_path,
            [
                ["姓名", "備註", "編號"],
                ["王小明", "王小明的身分證A123456789", 4595257],
            ],
        )

        response = _post_scrub_file(
            client,
            "名單.xlsx",
            content,
            entities=["PERSON", "TW_NATIONAL_ID"],
            column_strategies={"姓名": "mask_all", "備註": "ner", "編號": "skip"},
        )

        assert response.status_code == 200
        output = BytesIO(base64.b64decode(response.json()["file_b64"]))
        workbook = load_workbook(output, read_only=True)
        rows = list(workbook.active.iter_rows(values_only=True))
        workbook.close()
        assert rows[0] == ("姓名", "備註", "編號")
        assert rows[1] == ("王○○", "王○○的身分證A1******89", "4595257")
        # Excel 若已把前導 0 吃掉，來源只剩數值 4595257，程式無法可靠推回原始 8 碼；
        # 本驗收鎖定「整數轉字串且不產生 .0」，不臆測補零。

    def test_report_contains_only_masked_originals_and_locations(self, client) -> None:
        response = _post_scrub_file(
            client,
            "report.txt",
            "A123456789".encode("utf-8"),
            entities=["TW_NATIONAL_ID"],
        )

        assert response.status_code == 200
        report = response.json()["report"]
        datetime.fromisoformat(report["timestamp"])
        assert report["tool"] == "TW-PII-Scrubber"
        assert report["stats"] == {"TW_NATIONAL_ID": 1}
        assert "A123456789" not in json.dumps(report, ensure_ascii=False)
        assert report["findings"] == [
            {
                "entity_type": "TW_NATIONAL_ID",
                "original_masked": "A1******89",
                "score": 0.95,
                "recognizer": "TwNationalIdRecognizer",
                "replacement": "A1******89",
                "location": {"start": 0, "end": 10},
            }
        ]

    def test_missing_column_strategy_is_rejected(self, client) -> None:
        response = _post_scrub_file(
            client,
            "名單.csv",
            _csv_bytes([["姓名", "備註"], ["王小明", "內容"]]),
            column_strategies={"姓名": "mask_all"},
        )

        assert response.status_code == 400
        assert "備註" in response.json()["detail"]


class TestFileErrors:
    @pytest.mark.parametrize("endpoint", ["/api/scrub-file/preview", "/api/scrub-file"])
    def test_rejects_unsupported_extension(self, client, endpoint: str) -> None:
        response = client.post(
            endpoint,
            files={"file": ("名單.pdf", b"content", "application/pdf")},
        )

        assert response.status_code == 415

    @pytest.mark.parametrize("endpoint", ["/api/scrub-file/preview", "/api/scrub-file"])
    def test_rejects_undecodable_text(self, client, endpoint: str) -> None:
        response = client.post(
            endpoint,
            files={"file": ("broken.txt", b"\xff\xff", "text/plain")},
        )

        assert response.status_code == 400
        assert "請轉存為 UTF-8" in response.json()["detail"]

    @pytest.mark.parametrize("endpoint", ["/api/scrub-file/preview", "/api/scrub-file"])
    def test_rejects_files_larger_than_ten_mebibytes(
        self, client, endpoint: str
    ) -> None:
        content = b"a" * (10 * 1024 * 1024 + 1)

        response = client.post(
            endpoint,
            files={"file": ("oversize.txt", content, "text/plain")},
        )

        assert response.status_code == 413


class TestBundleZip:
    def test_returns_openable_zip_with_matching_contents(self, client) -> None:
        response = client.post(
            "/api/bundle-zip",
            json={
                "entries": [
                    {
                        "filename": "result.txt",
                        "content_b64": base64.b64encode("脫敏完成".encode()).decode(),
                    },
                    {
                        "filename": "批次報告.json",
                        "content_b64": base64.b64encode(b'{"ok": true}').decode(),
                    },
                ]
            },
        )

        assert response.status_code == 200
        body = response.json()
        assert body["filename"].startswith("tw-pii-scrubbed-")
        assert body["filename"].endswith(".zip")
        with ZipFile(BytesIO(base64.b64decode(body["zip_b64"]))) as archive:
            assert archive.namelist() == ["result.txt", "批次報告.json"]
            assert archive.read("result.txt") == "脫敏完成".encode()
            assert archive.read("批次報告.json") == b'{"ok": true}'

    def test_sanitizes_path_components(self, client) -> None:
        response = client.post(
            "/api/bundle-zip",
            json={
                "entries": [
                    {
                        "filename": "../../evil.txt",
                        "content_b64": base64.b64encode(b"safe").decode(),
                    }
                ]
            },
        )

        assert response.status_code == 200
        with ZipFile(BytesIO(base64.b64decode(response.json()["zip_b64"]))) as archive:
            assert archive.namelist() == ["evil.txt"]
            assert archive.read("evil.txt") == b"safe"

    def test_deduplicates_names_before_extension(self, client) -> None:
        encoded = base64.b64encode(b"same-name").decode()
        response = client.post(
            "/api/bundle-zip",
            json={
                "entries": [
                    {"filename": "report.csv", "content_b64": encoded},
                    {"filename": "report.csv", "content_b64": encoded},
                    {"filename": "report.csv", "content_b64": encoded},
                ]
            },
        )

        assert response.status_code == 200
        with ZipFile(BytesIO(base64.b64decode(response.json()["zip_b64"]))) as archive:
            assert archive.namelist() == [
                "report.csv",
                "report (2).csv",
                "report (3).csv",
            ]

    def test_rejects_more_than_max_entries(self, client) -> None:
        response = client.post(
            "/api/bundle-zip",
            json={
                "entries": [
                    {"filename": f"{index}.txt", "content_b64": ""}
                    for index in range(41)
                ]
            },
        )

        assert response.status_code == 400

    def test_rejects_invalid_base64(self, client) -> None:
        response = client.post(
            "/api/bundle-zip",
            json={"entries": [{"filename": "bad.txt", "content_b64": "%%%"}]},
        )

        assert response.status_code == 400

    def test_rejects_decoded_total_over_limit(self, client, monkeypatch) -> None:
        monkeypatch.setattr("app.config.MAX_ZIP_TOTAL_BYTES", 3)
        response = client.post(
            "/api/bundle-zip",
            json={
                "entries": [
                    {
                        "filename": "large.txt",
                        "content_b64": base64.b64encode(b"four").decode(),
                    }
                ]
            },
        )

        assert response.status_code == 413
